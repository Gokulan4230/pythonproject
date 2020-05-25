import openpyxl as xl
import pathlib
from openpyxl.chart import BarChart3D, Reference


def excelsheet_read_processor(excelfilename):
    path = pathlib.Path()
    address = path.glob(excelfilename)
    for file in address:
        print(file.exists())
    wb = xl.load_workbook(file)
    sheet = wb['Sheet1']
    for cell_value in range(2, sheet.max_row + 1):
        print(sheet.cell(cell_value, 2).value)
        annual_salary = sheet.cell(cell_value, 2).value * 12
        sheet.cell(cell_value, 3).value = annual_salary
    values = Reference(sheet,
                       min_row=2,
                       max_row=sheet.max_row,
                       min_col=3,
                       max_col=3,
                       )
    chart = BarChart3D()
    chart.add_data(values)
    sheet.add_chart(chart, 'D2')
    wb.save(excelfilename)


excelsheet_read_processor(excelfilename='exceldata.xlsx')
